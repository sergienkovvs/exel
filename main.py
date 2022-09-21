from openpyxl import Workbook

exel_file = Workbook()
exel_sheet = exel_file.create_sheet(title="CC 200+")
exel_sheet["A1"] = 1
exel_sheet["A2"] = "dssds"
exel_sheet["A3"] = "dfdfs"
exel_sheet["A4"] = "sdjnsv"
exel_sheet.cell(row=5, column=5).value = "sdsdf sdfsdf"
CC = (
    ("sdfsdf", "sdfsfsd"),
    ("afdsf", "asaffdssdf")
)
for row in CC:
    exel_sheet.append(row)
exel_file.save(filename="qwer.xlsx")
